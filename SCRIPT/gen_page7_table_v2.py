# -*- coding: utf-8 -*-
# page7 分组汇总表 v2（方案二：主表去安全/民生拆分，双轨「体检/诉求」表达；无渐变横条）。
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "DATA", "analysis", "page7小结", "page7_分组汇总_2026-08-14_v2.xlsx")

HDR_FILL = PatternFill("solid", fgColor="404040")
HDR_FONT = Font(bold=True, color="FFFFFF")
LAYER_FILL = {
    "双高": PatternFill("solid", fgColor="FCE4EC"),     # 浅红
    "客观高": PatternFill("solid", fgColor="E3EDF7"),   # 浅蓝
    "主观高": PatternFill("solid", fgColor="FDEBD9"),   # 浅橙
}
LAYER_HEAD = {
    "双高": PatternFill("solid", fgColor="C00000"),     # 深红
    "客观高": PatternFill("solid", fgColor="1F4E79"),   # 深蓝
    "主观高": PatternFill("solid", fgColor="C55A11"),   # 深橙
}
LABEL_FONT = {
    "双高": Font(bold=True, color="C00000"),
    "客观高": Font(bold=True, color="1F4E79"),
    "主观高": Font(bold=True, color="C55A11"),
}
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# 20 社区：名称, 楼栋, 体检点(安全+民生), 诉求件(安全+民生), 层
ROWS = [
    ("营盘路社区", 66, 37, 117, "双高"),
    ("宝联社区", 94, 29, 156, "双高"),
    ("汕头路社区", 39, 17, 107, "双高"),
    ("胜利四路社区", 44, 16, 115, "双高"),
    ("胜利二路社区", 30, 9, 111, "双高"),
    ("深圳路社区", 127, 107, 7, "客观高"),
    ("西峡社区", 78, 91, 35, "客观高"),
    ("金安岭社区", 56, 84, 20, "客观高"),
    ("镇境山社区", 78, 82, 5, "客观高"),
    ("幸福路社区", 58, 65, 1, "客观高"),
    ("新隆康路社区", 130, 62, 64, "客观高"),
    ("果园路社区", 93, 51, 73, "客观高"),
    ("桥北社区", 48, 37, 0, "客观高"),
    ("朝阳路社区", 84, 11, 594, "主观高"),
    ("万达社区", 69, 7, 501, "主观高"),
    ("港务社区", 106, 17, 417, "主观高"),
    ("建设社区", 66, 9, 372, "主观高"),
    ("岳湾路社区", 83, 2, 225, "主观高"),
    ("大学路社区", 97, 7, 222, "主观高"),
    ("伍临路社区", 55, 8, 221, "主观高"),
]


def _dens(n, b):
    return round(n / b * 100, 1) if b else 0


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "page7分组汇总"

    headers = ["序", "社区", "楼栋", "体检问题（点）", "市民诉求（件）", "每百栋（体检/诉求）", "综合评估"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER

    layer_order = ["双高", "客观高", "主观高"]
    layer_titles = {
        "双高": "① 双高 · 体检+诉求都高（5 个）",
        "客观高": "② 客观隐患高 · 体检独证（诉求未暴露 · 8 个）",
        "主观高": "③ 主观诉求高 · 热线独证（体检未印证 · 7 个）",
    }
    layer_labels = {
        "双高": "双高 ★",
        "客观高": "客观隐患高\n诉求未暴露",
        "主观高": "主观诉求高\n体检未印证",
    }

    r = 2
    seq = 0
    for layer in layer_order:
        # 层小标题行
        ws.cell(r, 1, layer_titles[layer])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        for c in range(1, 8):
            cell = ws.cell(r, c)
            cell.fill = LAYER_HEAD[layer]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = BORDER
        ws.cell(r, 1).alignment = LEFT
        r += 1
        for name, b, tj, sub, _l in ROWS:
            if _l != layer:
                continue
            seq += 1
            d_tj = _dens(tj, b)
            d_sub = _dens(sub, b)
            dens_str = f"{d_tj} / {d_sub if sub else '—'}"
            sub_str = sub if sub else "—"
            vals = [seq, name, b, tj, sub_str, dens_str, layer_labels[layer]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = BORDER
                cell.fill = LAYER_FILL[layer]
                cell.alignment = CENTER if c in (1, 3, 4, 5, 6) else LEFT
                if c == 7:
                    cell.font = LABEL_FONT[layer]
                    cell.alignment = CENTER
            r += 1

    # 表尾口径注
    r += 1
    notes = [
        "排序：双高 → 客观高 → 主观高；层内按绝对量降序（体检点 / 诉求件），密度仅作资格与旁证、不进排序键。",
        "体检问题（点）= 体检安全 + 体检民生（安全/民生拆分见备查 sheet）；市民诉求（件）= 12345 九类（安全 4 + 民生 5，剔「其他」）。",
        "每百栋 = 问题数 ÷ 楼栋数 × 100（体检/诉求 两列）；「—」= 无隐患/无诉求（体检未标定该方面问题）。",
        "分层阈值 = 客观/主观密度各自全样本 p81（前 19%）：客观 ≥28.57 点/百栋、主观 ≥146.67 件/百栋。",
        "港务（诉求 417 件·全市第 3）落主观高层，属「诉求型」非「隐患型」——体检优先是诊断序、任务量是实施序。",
    ]
    for n in notes:
        ws.cell(r, 1, n)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 1).font = Font(size=9, color="666666")
        r += 1

    widths = [5, 14, 7, 12, 12, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 备查 sheet：安全/民生拆分
    ws2 = wb.create_sheet("安全民生拆分备查")
    hdr2 = ["社区", "体检安全", "体检民生", "诉求安全", "诉求民生", "体检点", "诉求件", "层"]
    for c, h in enumerate(hdr2, 1):
        cell = ws2.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER
    # 安全/民生拆分值（从分组汇总原表）
    detail = [
        ("营盘路社区", 0, 37, 10, 107, 37, 117, "双高"),
        ("宝联社区", 3, 26, 29, 127, 29, 156, "双高"),
        ("汕头路社区", 14, 3, 10, 97, 17, 107, "双高"),
        ("胜利四路社区", 6, 10, 20, 95, 16, 115, "双高"),
        ("胜利二路社区", 6, 3, 9, 102, 9, 111, "双高"),
        ("深圳路社区", 78, 29, 3, 4, 107, 7, "客观高"),
        ("西峡社区", 82, 9, 3, 32, 91, 35, "客观高"),
        ("金安岭社区", 73, 11, 9, 11, 84, 20, "客观高"),
        ("镇境山社区", 67, 15, 4, 1, 82, 5, "客观高"),
        ("幸福路社区", 48, 17, 1, 0, 65, 1, "客观高"),
        ("新隆康路社区", 47, 15, 6, 58, 62, 64, "客观高"),
        ("果园路社区", 25, 26, 11, 62, 51, 73, "客观高"),
        ("桥北社区", 17, 20, 0, 0, 37, 0, "客观高"),
        ("朝阳路社区", 7, 4, 58, 536, 11, 594, "主观高"),
        ("万达社区", 0, 7, 50, 451, 7, 501, "主观高"),
        ("港务社区", 2, 15, 83, 334, 17, 417, "主观高"),
        ("建设社区", 0, 9, 52, 320, 9, 372, "主观高"),
        ("岳湾路社区", 0, 2, 19, 206, 2, 225, "主观高"),
        ("大学路社区", 1, 6, 44, 178, 7, 222, "主观高"),
        ("伍临路社区", 0, 8, 20, 201, 8, 221, "主观高"),
    ]
    r2 = 2
    for row in detail:
        for c, v in enumerate(row, 1):
            cell = ws2.cell(r2, c, v)
            cell.border = BORDER
            cell.alignment = CENTER if c > 1 else LEFT
            if c == 8:
                cell.font = LABEL_FONT[row[-1]]
        r2 += 1
    for i, w in enumerate([14, 10, 10, 10, 10, 8, 8, 10], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(OUT)
    print("[OK]", os.path.basename(OUT))


if __name__ == "__main__":
    main()
