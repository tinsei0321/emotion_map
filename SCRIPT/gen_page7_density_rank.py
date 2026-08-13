# -*- coding: utf-8 -*-
"""page7 汇总表 v2：安全组/民生组 双层表头 + 实心数据条（#1F4E79 体检 / #C55A11 诉求）。"""
import csv, pathlib, sys
import re, shutil, zipfile
from collections import Counter
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import DataBarRule

BASE = pathlib.Path("DATA/analysis")

def read_csv(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))

def num(s):
    try:
        return float(s)
    except Exception:
        return 0.0

safe = read_csv(BASE / "安全韧性" / "安全韧性_社区3类矩阵.csv")
mins = read_csv(BASE / "民生基础" / "民生_社区5类矩阵.csv")
r123 = read_csv(BASE / "12345主观" / "12345_社区x9类_西陵伍家.csv")
denom = read_csv(BASE / "page7小结" / "社区规模分母_174.csv")

safe_map = {r["社区"]: r for r in safe}
mins_map = {r["社区"]: r for r in mins}
r123_map = {r["社区"]: r for r in r123}
denom_map = {r["社区"]: r for r in denom}

SAFE_COLS = ["管网安全", "出行安全", "消防安全", "环境安全"]
MIN_COLS = ["噪声", "停车", "住宅", "出行", "物业"]

rows = []
for c, d in denom_map.items():
    bldg = num(d["bldg_n"])
    if bldg <= 0:
        continue
    tj_safe = num(safe_map[c]["总点数"]) if c in safe_map else 0.0
    tj_min = num(mins_map[c]["总点数"]) if c in mins_map else 0.0
    r_safe = sum(num(r123_map[c][k]) for k in SAFE_COLS) if c in r123_map else 0.0
    r_min = sum(num(r123_map[c][k]) for k in MIN_COLS) if c in r123_map else 0.0
    obj = tj_safe + tj_min
    sub = r_safe + r_min
    rows.append(dict(
        社区=c, 楼栋=bldg,
        体检安全=tj_safe, 体检民生=tj_min,
        热线安全=r_safe, 热线民生=r_min,
        客观=obj, 主观=sub, 任务量=obj + sub,
        体检安全密度=tj_safe / bldg * 100, 体检民生密度=tj_min / bldg * 100,
        热线安全密度=r_safe / bldg * 100, 热线民生密度=r_min / bldg * 100,
        客观密度=obj / bldg * 100, 主观密度=sub / bldg * 100,
    ))

active = [r for r in rows if r["客观"] > 0 or r["主观"] > 0]

def pct(xs, p):
    s = sorted(xs)
    if not s:
        return 0.0
    k = (len(s) - 1) * p
    i = int(k)
    f = k - i
    return s[i] * (1 - f) + s[min(i + 1, len(s) - 1)] * f

obj_p75 = pct([r["客观密度"] for r in active], 0.75)
sub_p75 = pct([r["主观密度"] for r in active], 0.75)

for r in rows:
    oh = r["客观密度"] >= obj_p75 and r["客观"] > 0
    sh = r["主观密度"] >= sub_p75 and r["主观"] > 0
    if oh and sh:
        r["分层"] = "双高"
    elif oh:
        r["分层"] = "问题指标高"
    elif sh:
        r["分层"] = "诉求声量高"
    else:
        r["分层"] = "其余"
    r["低置信"] = "低置信" if r["楼栋"] < 20 else ""

shuang = sorted([r for r in rows if r["分层"] == "双高"], key=lambda r: -r["客观"])
ke = sorted([r for r in rows if r["分层"] == "问题指标高"], key=lambda r: -r["客观"])
zh = sorted([r for r in rows if r["分层"] == "诉求声量高"], key=lambda r: -r["主观"])
top = shuang + ke[:8] + zh[:7]

LABELS = {"双高": "双高", "问题指标高": "客观隐患高·诉求未暴露", "诉求声量高": "主观诉求高·体检未印证"}

# ---- 写 Excel ----
wb = Workbook()
ws = wb.active
ws.title = "page7分组汇总"

# 双层表头
ws["A1"] = "序号"; ws["B1"] = "名称"; ws["C1"] = "楼栋"
ws["D1"] = "安全韧性问题"; ws["F1"] = "民生基础需求"
ws["H1"] = "问题指标"; ws["I1"] = "群众诉求"; ws["J1"] = "评估"; ws["K1"] = "备注"
ws["D2"] = "体检"; ws["E2"] = "诉求"; ws["F2"] = "体检"; ws["G2"] = "诉求"

ws.merge_cells("D1:E1")
ws.merge_cells("F1:G1")
for col in ("A", "B", "C", "H", "I", "J", "K"):
    ws.merge_cells(f"{col}1:{col}2")

for i, r in enumerate(top, start=3):
    ws.cell(row=i, column=1, value=i - 2)
    ws.cell(row=i, column=2, value=r["社区"])
    ws.cell(row=i, column=3, value=int(r["楼栋"]))
    ws.cell(row=i, column=4, value=round(r["体检安全密度"], 1))
    ws.cell(row=i, column=5, value=round(r["热线安全密度"], 1))
    ws.cell(row=i, column=6, value=round(r["体检民生密度"], 1))
    ws.cell(row=i, column=7, value=round(r["热线民生密度"], 1))
    ws.cell(row=i, column=8, value=int(r["客观"]))
    ws.cell(row=i, column=9, value=int(r["主观"]))
    ws.cell(row=i, column=10, value=LABELS[r["分层"]])
    ws.cell(row=i, column=11, value=r["低置信"])

last = 2 + len(top)
# 体检 = 蓝 #1F4E79，诉求 = 橙 #C55A11，实心填充（去渐变）
rules = [
    ("D3:D" + str(last), "1F4E79"),
    ("F3:F" + str(last), "1F4E79"),
    ("E3:E" + str(last), "C55A11"),
    ("G3:G" + str(last), "C55A11"),
]
for rng, color in rules:
    rule = DataBarRule(start_type="num", start_value=0, end_type="num",
                       end_value=700, color=color, showValue=True)
    rule.dataBar.gradient = False
    ws.conditional_formatting.add(rng, rule)

thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
header_fill = PatternFill("solid", fgColor="EDEDED")

for row in ws.iter_rows(min_row=1, max_row=last):
    for cell in row:
        cell.border = border
        cell.alignment = center
for row in (1, 2):
    for cell in ws[row]:
        cell.font = bold
        cell.fill = header_fill

widths = {"A": 6, "B": 13, "C": 7, "D": 11, "E": 11, "F": 11, "G": 11, "H": 10, "I": 10, "J": 13, "K": 9}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 18

# 表尾说明
note1 = last + 2
note2 = last + 3
ws.cell(row=note1, column=1, value="排序：双高 → 客观隐患高（按体检点降序）→ 主观诉求高（按诉求件数降序）；密度仅作资格与旁证，不作排序键。")
ws.cell(row=note1, column=1).font = Font(italic=True, color="808080")
ws.cell(row=note2, column=1, value="脚注：体检高 = 客观隐患优先整治；诉求高 = 主观诉求集中须回应。两者都是行动对象、同属高优先，只是处置逻辑不同：一处隐患、一应诉求。")
ws.cell(row=note2, column=1).font = Font(italic=True, color="808080")

out = BASE / "page7小结" / "page7_分组汇总_2026-08-14.xlsx"
wb.save(out)

# 后处理：openpyxl 不暴露 dataBar 的 gradient 属性，手动补 gradient="0"（实心填充·去渐变）
def set_solid_bars(path):
    tmp = str(path) + ".tmp"
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml") and "worksheet" in item.filename:
                txt = data.decode("utf-8")
                txt = re.sub(r'<dataBar(?![^>]*gradient)([^>]*)>', r'<dataBar\1 gradient="0">', txt)
                data = txt.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp, path)

set_solid_bars(out)

print("已写出:", out)
print("双高:", len(shuang), " 客观隐患高:", len(ke), " 主观诉求高:", len(zh))
for i, r in enumerate(top, 1):
    print(f"{i:2d} [{LABELS[r['分层']]}] {r['社区']:<8} 楼栋{int(r['楼栋']):>3} 体检点{int(r['客观']):>3} 诉求件{int(r['主观']):>3}")
