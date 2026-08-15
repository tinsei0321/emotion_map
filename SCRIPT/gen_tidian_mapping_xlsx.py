# -*- coding: utf-8 -*-
# 生成「体检点数据类型来源对应」Excel（放 DATA/analysis/，供查阅）。
# 含「3268→928+757」提取漏斗：原始 3268 = 提取 1685（安全928+民生757）+ 排除 1583（附加56+设施点55+覆盖率1472）。
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "DATA", "analysis", "体检点数据类型来源对应_2026-08-14.xlsx")

HDR_FILL = PatternFill("solid", fgColor="404040")
HDR_FONT = Font(bold=True, color="FFFFFF")
EXCL_FILL = PatternFill("solid", fgColor="F2F2F2")   # 排除行浅灰
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")    # 小计/汇总行浅蓝
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")   # 提取行浅绿（可选，不填也清晰）
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write(ws, r, vals, fill=None):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.border = BORDER
        cell.alignment = LEFT
        if fill:
            cell.fill = fill


def _header(ws, r, vals):
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER


def build_detail(ws):
    _header(ws, 1, ["源头维度", "源头子类", "源头图斑类型", "图斑数", "提取/排除", "提取去向（8类）", "方面", "备注"])
    rows = [
        # 维度, 子类, 类型, 数量, 提取/排除, 去向, 方面, 备注
        ("住房", "安全耐久", "结构隐患住宅", 42, "提取", "安全·住房", "安全韧性", ""),
        ("住房", "安全耐久", "燃气隐患住宅", 6, "提取", "安全·市政管网", "安全韧性", "燃气归管网"),
        ("住房", "安全耐久", "楼道隐患住宅", 240, "提取", "安全·安全消防", "安全韧性", ""),
        ("住房", "安全耐久", "围护隐患住宅", 454, "提取", "安全·住房", "安全韧性", ""),
        ("住房", "功能完备", "非成套住宅", 31, "提取", "民生·住房", "民生基础", ""),
        ("住房", "功能完备", "管线管道破损住宅", 186, "提取", "安全·市政管网", "安全韧性", "管线归管网"),
        ("住房", "功能完备", "适老化改造住宅", 39, "提取", "民生·住房", "民生基础", ""),
        ("住房", "绿色智能", "数字化改造住宅", 25, "排除·附加", "—", "其他", "绿色智能=附加发展类"),
        ("小区", "设施完善", "养老未达标小区", 2, "提取", "民生·公服设施", "民生基础", ""),
        ("小区", "设施完善", "婴幼儿照护未达标", 34, "提取", "民生·公服设施", "民生基础", ""),
        ("小区", "设施完善", "幼儿园未达标", 11, "提取", "民生·公服设施", "民生基础", ""),
        ("小区", "设施完善", "小学学位缺口", 31, "提取", "民生·公服设施", "民生基础", ""),
        ("小区", "设施完善", "停车泊位缺口", 140, "提取", "民生·停车设施", "民生基础", ""),
        ("小区", "设施完善", "充电桩缺口", 84, "提取", "民生·停车设施", "民生基础", ""),
        ("小区", "设施完善", "电动车充电未配建", 50, "提取", "民生·停车设施", "民生基础", ""),
        ("小区", "环境宜居", "步行道不达标", 24, "提取", "民生·交通设施", "民生基础", ""),
        ("小区", "环境宜居", "公共活动场地未达标", 27, "提取", "民生·公服设施", "民生基础", ""),
        ("小区", "管理健全", "未实施物业小区", 273, "提取", "民生·物业街面", "民生基础", ""),
        ("小区", "管理健全", "智慧化改造小区", 31, "排除·附加", "—", "其他", "智慧化=附加发展类"),
        ("街区", "功能完善", "中学（设施点）", 21, "排除·设施点", "—", "—", "设施数量·非问题"),
        ("街区", "功能完善", "中学覆盖率", 535, "排除·覆盖率", "—", "—", "覆盖率=面积指标"),
        ("街区", "功能完善", "公园覆盖率", 515, "排除·覆盖率", "—", "—", "覆盖率=面积指标"),
        ("街区", "功能完善", "菜市场（设施点）", 34, "排除·设施点", "—", "—", "设施数量·非问题"),
        ("街区", "功能完善", "菜市场覆盖率", 422, "排除·覆盖率", "—", "—", "覆盖率=面积指标"),
        ("街区", "功能完善", "多功能运动场地未达标", 6, "提取", "民生·公服设施", "民生基础", ""),
        ("街区", "整洁有序", "乱停乱放车辆道路", 5, "提取", "民生·物业街面", "民生基础", "线转点"),
    ]
    r = 2
    for row in rows:
        excl = str(row[4]).startswith("排除")
        _write(ws, r, list(row), fill=EXCL_FILL if excl else None)
        r += 1
    # 汇总行
    total = sum(row[3] for row in rows)
    extract = sum(row[3] for row in rows if str(row[4]) == "提取")
    exclude = total - extract
    _write(ws, r, ["合计", "", f"26 图层 / {len(rows)} 类图斑", total, f"提取 {extract} / 排除 {exclude}", "", "", "3268 = 提取1685 + 排除1583"], fill=SUB_FILL)
    widths = [8, 10, 22, 8, 12, 16, 10, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def build_summary(ws):
    _header(ws, 1, ["最终社区点（8类）", "方面", "源图斑构成", "源图斑数", "提取点数据", "社区矩阵", "差异", "差异说明"])
    rows = [
        ("安全·住房", "安全韧性", "结构隐患42 + 围护隐患454", 496, 496, 496, 0, "零丢失"),
        ("安全·安全消防", "安全韧性", "楼道隐患240", 240, 240, 240, 0, "零丢失"),
        ("安全·市政管网", "安全韧性", "管线破损186 + 燃气隐患6", 192, 192, 192, 0, "零丢失"),
        ("安全小计", "—", "—", 928, 928, 928, 0, "零丢失"),
        ("民生·公服设施", "民生基础", "养老2+婴幼儿34+幼儿园11+学位31+活动场地27+运动场地6", 111, 111, 111, 0, "2点范围外忽略"),
        ("民生·住房", "民生基础", "非成套31 + 适老化39", 70, 70, 70, 0, "零丢失"),
        ("民生·停车设施", "民生基础", "泊位140+充电桩84+电动车50", 274, 274, 274, 0, "零丢失"),
        ("民生·交通设施", "民生基础", "不达标步行道24", 24, 24, 24, 0, "零丢失"),
        ("民生·物业街面", "民生基础", "未物业273 + 乱停乱放5", 278, 278, 278, 0, "零丢失"),
        ("民生小计", "—", "—", 757, 757, 757, 0, "2点范围外忽略"),
    ]
    r = 2
    for row in rows:
        subtotal = "小计" in str(row[0])
        _write(ws, r, list(row), fill=SUB_FILL if subtotal else None)
        r += 1
    widths = [18, 10, 42, 9, 10, 10, 8, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def build_funnel(ws):
    _header(ws, 1, ["环节", "分类", "数量", "说明"])
    rows = [
        ("原始数据库", "总图斑（26 图层）", 3268, "住房/小区/街区三维度全部要素"),
        ("", "提取·安全韧性", 928, "结构42+围护454+楼道240+燃气6+管线186"),
        ("", "提取·民生基础", 757, "养老2+婴幼儿34+幼儿园11+学位31+停车140+充电桩84+电动车50+步行道24+活动场地27+未物业273+运动场地6+乱停乱放5+非成套31+适老化39"),
        ("", "提取小计", 1685, "928 + 757 = 负面图斑全提取"),
        ("", "排除·附加", 56, "数字化25 + 智慧化31（绿色智能=附加发展类）"),
        ("", "排除·设施点", 55, "中学设施点21 + 菜市场设施点34（设施数量·非问题）"),
        ("", "排除·覆盖率", 1472, "中学覆盖率535 + 公园覆盖率515 + 菜市场覆盖率422（面积指标·点数多=覆盖好）"),
        ("", "排除小计", 1583, "56 + 55 + 1472"),
        ("校验", "3268 = 提取 + 排除", 3268, "1685 + 1583 = 3268 ✓"),
    ]
    r = 2
    for row in rows:
        is_total = "3268" in str(row[2]) and row[0] in ("原始数据库", "校验")
        is_sub = "小计" in str(row[1])
        fill = SUB_FILL if (is_total or is_sub) else None
        _write(ws, r, list(row), fill=fill)
        r += 1
    widths = [10, 20, 10, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def main():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "提取漏斗"
    build_funnel(ws1)
    ws2 = wb.create_sheet("类型来源对应明细")
    build_detail(ws2)
    ws3 = wb.create_sheet("汇总对照")
    build_summary(ws3)
    wb.save(OUT)
    print("[OK]", os.path.basename(OUT))


if __name__ == "__main__":
    main()
