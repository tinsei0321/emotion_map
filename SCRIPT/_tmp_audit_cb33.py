# -*- coding: utf-8 -*-
# CB-33 评估方只读核验（不写交付物·不 git）· 跑完即删
# 独立复算：inside 范围 / 分方面 / 类9 / 社区村TOP / 交付物feature数 / Excel实际值
import json, os, sys
from collections import Counter
import pandas as pd
from shapely.geometry import shape
from shapely.ops import unary_union
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ANALYSIS = os.path.join(ROOT, "DATA", "analysis")
SUBJ = os.path.join(ANALYSIS, "12345主观")

def load(p):
    with open(p, encoding="utf-8") as f:
        return json.load(f)

rng = load(os.path.join(ANALYSIS, "12345_西陵+伍家岗.geojson"))
union = unary_union([shape(f["geometry"]) for f in rng["features"]])
pts = load(os.path.join(SUBJ, "12345_有坐标点.geojson"))
feats = pts["features"]
def in_union(g):
    return union.contains(g) or union.covers(g) or union.touches(g)
inside = [f for f in feats if in_union(shape(f["geometry"]))]
print(f"[复算] 总有坐标点 {len(feats)} | 命中西陵+伍家 {len(inside)}")

DECL = {"安全韧性": 2041, "民生基础": 11348}
DECL_C9 = {
    "安全韧性": {"管网安全": 1102, "出行安全": 723, "消防安全": 198, "环境安全": 18},
    "民生基础": {"噪声": 3689, "停车": 1870, "住宅": 2466, "出行": 1027, "物业": 2296},
}
for asp in ("安全韧性", "民生基础"):
    sub = [f for f in inside if f["properties"].get("方面") == asp]
    ok_c = [f for f in sub if f["properties"].get("geocode_status") == "ok" and f["properties"].get("社区")]
    ok_v = [f for f in sub if f["properties"].get("geocode_status") == "ok" and f["properties"].get("村")]
    region = [f for f in sub if f["properties"].get("geocode_status") == "region"]
    out = [f for f in sub if f["properties"].get("geocode_status") == "ok" and not f["properties"].get("社区") and not f["properties"].get("村")]
    c9 = Counter(f["properties"].get("类9") for f in sub)
    print(f"\n=== {asp} ===")
    print(f"  复算总 {len(sub)} (声明 {DECL[asp]}) = 社区{len(ok_c)} 村{len(ok_v)} 区级{len(region)} 范围外{len(out)}  合计校验{len(ok_c)+len(ok_v)+len(region)+len(out)}")
    print(f"  类9复算: {dict(sorted(c9.items(), key=lambda x: -x[1]))}")
    print(f"  类9声明: {DECL_C9[asp]}")
    # 交付物点文件 feature 数
    dp = load(os.path.join(SUBJ, f"12345_{asp}_西陵伍家点.geojson"))
    print(f"  交付物点文件 {len(dp['features'])} (声明 {DECL[asp]}) {'OK' if len(dp['features']) == DECL[asp] else 'MISMATCH'}")
    # 社区 TOP10
    cc = Counter(f["properties"]["社区"] for f in ok_c)
    tot = len(ok_c)
    print(f"  社区TOP10 (精确社区口径 n={tot}):")
    for i, (n, cnt) in enumerate(cc.most_common(10), 1):
        print(f"    {i}. {n} {cnt} ({cnt/tot*100:.1f}%)")
    vc = Counter(f["properties"]["村"] for f in ok_v)
    vt = len(ok_v)
    print(f"  村排行 (精确村口径 n={vt}):")
    for i, (n, cnt) in enumerate(vc.most_common(10), 1):
        print(f"    {i}. {n} {cnt} ({cnt/vt*100:.1f}%)")

cm = pd.read_csv(os.path.join(SUBJ, "12345_社区x9类_西陵伍家.csv"))
vl = pd.read_csv(os.path.join(SUBJ, "12345_村x9类_西陵伍家.csv"))
print(f"\n[csv] 社区 {len(cm)} 行 · 村 {len(vl)} 行")

# Excel 实际落盘值
wb = openpyxl.load_workbook(os.path.join(ANALYSIS, "图数表出图_PPT表格汇总.xlsx"))
for sh in ("page4", "page6"):
    ws = wb[sh]
    print(f"\n[Excel {sh}]")
    print(f"  A3={ws['A3'].value}")
    # 表A 件数 B5:B9 / 占比 C / 社区数 D
    rows_a = 5 if sh == "page4" else 5
    n_a = 4 if sh == "page4" else 5
    vals = []
    for r in range(rows_a, rows_a + n_a):
        vals.append((ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value))
    print(f"  表A(件数/占比/社区数): {vals}")
    # 表B 合计行
    if sh == "page4":
        hb = 16
    else:
        hb = 18
    print(f"  表B合计行{hb}: 社区{ws.cell(hb,2).value} 村{ws.cell(hb,3).value} 区级+外{ws.cell(hb,4).value} 命中合计{ws.cell(hb,5).value} {ws.cell(hb,6).value}")
