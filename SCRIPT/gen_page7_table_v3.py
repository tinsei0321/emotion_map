# -*- coding: utf-8 -*-
# page7 分组汇总表 v3（横条对比版：体检/诉求合并、实体填充横条·无渐变）。
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "DATA", "analysis", "page7小结", "page7_分组汇总_2026-08-14_v3横条.xlsx")

BLOCK = "█"
MAX_BAR = 22

HDR_FILL = PatternFill("solid", fgColor="404040")
HDR_FONT = Font(bold=True, color="FFFFFF")
LAYER_FILL = {
    "双高": PatternFill("solid", fgColor="FCE4EC"),
    "客观高": PatternFill("solid", fgColor="E3EDF7"),
    "主观高": PatternFill("solid", fgColor="FDEBD9"),
}
LAYER_HEAD = {
    "双高": PatternFill("solid", fgColor="C00000"),
    "客观高": PatternFill("solid", fgColor="1F4E79"),
    "主观高": PatternFill("solid", fgColor="C55A11"),
}
LABEL_FONT = {
    "双高": Font(bold=True, color="C00000"),
    "客观高": Font(bold=True, color="1F4E79"),
    "主观高": Font(bold=True, color="C55A11"),
}
TJ_FONT = Font(color="1F4E79", name="Consolas")
SUB_FONT = Font(color="C55A11", name="Consolas")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 名称, 楼栋, 体检点, 诉求件, 层
ROWS = [
    ("营盘路社区", 66, 37, 117, "双高"),
    ("宝联社区", 94, 29, 156, "双高"),
    ("汕头路社区", 39, 17, 107, "双高"),
    ("胜利四路社区", 44, 16, 115, "双高"),
    ("胜利二路社区", 30, 9, 111, "双高"),
    ("深圳路社区", 127, 107, 7, "客观高"),
    ("西峡社区", 78, 91, 35, "客观高"),
    ("金安岭社区", 56, 84, 20, "客观高"),
    ("镇境山社区", 78, 82, 5, "客观高"),
    ("幸福路社区", 58, 65, 1, "客观高"),
    ("新隆康路社区", 130, 62, 64, "客观高"),
    ("果园路社区", 93, 51, 73, "客观高"),
    ("桥北社区", 48, 37, 0, "客观高"),
    ("朝阳路社区", 84, 11, 594, "主观高"),
    ("万达社区", 69, 7, 501, "主观高"),
    ("港务社区", 106, 17, 417, "主观高"),
    ("建设社区", 66, 9, 372, "主观高"),
    ("岳湾路社区", 83, 2, 225, "主观高"),
    ("大学路社区", 97, 7, 222, "主观高"),
    ("伍临路社区", 55, 8, 221, "主观高"),
]

TJ_MAX = 150.0    # 体检密度满格（金安岭 150/百栋）
SUB_MAX = 726.0   # 诉求密度满格（万达 726/百栋）


def bar(n, maxv, val, num):
    """实体横条：长度=密度/满格×MAX_BAR，条后跟绝对量数字。val=0 用「—」。"""
    if val <= 0 or num <= 0:
        return "—"
    k = round(val / maxv * MAX_BAR)
    k = max(1, min(MAX_BAR, k))
    return BLOCK * k + " " + str(num)


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "page7分组汇总"

    headers = ["序", "社区", "楼栋", "体检（点·每百栋）", "诉求（件·每百栋）", "综合评估"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER

    layer_order = ["双高", "客观高", "主观高"]
    layer_titles = {
        "双高": "① 双高 · 体检+诉求都高（5 个）",
        "客观高": "② 客观隐患高 · 体检独证（8 个）",
        "主观高": "③ 主观诉求高 · 热线独证（7 个）",
    }
    layer_labels = {
        "双高": "双高 ★",
        "客观高": "客观隐患高\n诉求未暴露",
        "主观高": "主观诉求高\n体检未印证",
    }

    r = 2
    seq = 0
    for layer in layer_order:
        ws.cell(r, 1, layer_titles[layer])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        for c in range(1, 7):
            cell = ws.cell(r, c)
            cell.fill = LAYER_HEAD[layer]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = BORDER
        ws.cell(r, 1).alignment = LEFT
        r += 1
        for name, b, tj, sub, _l in ROWS:
            if _l != layer:
                continue
            seq += 1
            tj_den = tj / b * 100
            sub_den = sub / b * 100
            tj_cell = bar(1, TJ_MAX, tj_den, tj)
            sub_cell = bar(1, SUB_MAX, sub_den, sub)
            vals = [seq, name, b, tj_cell, sub_cell, layer_labels[layer]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = BORDER
                cell.fill = LAYER_FILL[layer]
                cell.alignment = CENTER if c in (1, 3, 6) else LEFT
                if c == 4:
                    cell.font = TJ_FONT
                elif c == 5:
                    cell.font = SUB_FONT
                elif c == 6:
                    cell.font = LABEL_FONT[layer]
                    cell.alignment = CENTER
            r += 1

    r += 1
    notes = [
        "横条 = 每百栋密度（体检满格 150 点/百栋、诉求满格 726 件/百栋，两轨分刻度）；条后数字 = 绝对量（体检点 / 诉求件）。",
        "体检 = 体检安全 + 体检民生（合并）；诉求 = 12345 九类（安全 4 + 民生 5，剔「其他」）；「—」= 无隐患/无诉求。",
        "排序：双高 → 客观高 → 主观高；层内按绝对量降序；分层阈值 = 密度 p81（客观 28.57 / 主观 146.67）。",
        "港务诉求 417 件（全市第 3）落主观高层，属「诉求型」非「隐患型」——体检优先是诊断序、任务量是实施序。",
    ]
    for n in notes:
        ws.cell(r, 1, n)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 1).font = Font(size=9, color="666666")
        r += 1

    widths = [5, 13, 7, 26, 26, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(OUT)
    print("[OK]", os.path.basename(OUT))


if __name__ == "__main__":
    main()
