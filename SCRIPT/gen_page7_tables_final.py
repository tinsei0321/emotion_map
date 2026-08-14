# -*- coding: utf-8 -*-
# page7 最终三张表（一张 Excel·三个 sheet·数据统一派生自同一数据源）。
# Sheet1 原表（安全/民生 × 体检/诉求 每百栋）· Sheet2 分类表（双高/客观高/主观高）· Sheet3 体检-诉求对比表（横条）。
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "DATA", "analysis", "page7小结", "page7_分组汇总_2026-08-14_最终.xlsx")

BLOCK = "█"
MAX_BAR = 22
TJ_MAX = 150.0
SUB_MAX = 726.0

HDR_FILL = PatternFill("solid", fgColor="404040")
HDR_FONT = Font(bold=True, color="FFFFFF")
LAYER_FILL = {
    "双高": PatternFill("solid", fgColor="FCE4EC"),
    "客观高": PatternFill("solid", fgColor="E3EDF7"),
    "主观高": PatternFill("solid", fgColor="FDEBD9"),
}
LAYER_HEAD = {
    "双高": PatternFill("solid", fgColor="C00000"),
    "客观高": PatternFill("solid", fgColor="1F4E79"),
    "主观高": PatternFill("solid", fgColor="C55A11"),
}
LABEL_FONT = {
    "双高": Font(bold=True, color="C00000"),
    "客观高": Font(bold=True, color="1F4E79"),
    "主观高": Font(bold=True, color="C55A11"),
}
TJ_FONT = Font(color="1F4E79", name="Consolas")
SUB_FONT = Font(color="C55A11", name="Consolas")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 唯一数据源：社区, 楼栋, 体检安全, 体检民生, 诉求安全, 诉求民生, 层
ROWS = [
    ("营盘路社区", 66, 0, 37, 10, 107, "双高"),
    ("宝联社区", 94, 3, 26, 29, 127, "双高"),
    ("汕头路社区", 39, 14, 3, 10, 97, "双高"),
    ("胜利四路社区", 44, 6, 10, 20, 95, "双高"),
    ("胜利二路社区", 30, 6, 3, 9, 102, "双高"),
    ("深圳路社区", 127, 78, 29, 3, 4, "客观高"),
    ("西峡社区", 78, 82, 9, 3, 32, "客观高"),
    ("金安岭社区", 56, 73, 11, 9, 11, "客观高"),
    ("镇境山社区", 78, 67, 15, 4, 1, "客观高"),
    ("幸福路社区", 58, 48, 17, 1, 0, "客观高"),
    ("新隆康路社区", 130, 47, 15, 6, 58, "客观高"),
    ("果园路社区", 93, 25, 26, 11, 62, "客观高"),
    ("桥北社区", 48, 17, 20, 0, 0, "客观高"),
    ("朝阳路社区", 84, 7, 4, 58, 536, "主观高"),
    ("万达社区", 69, 0, 7, 50, 451, "主观高"),
    ("港务社区", 106, 2, 15, 83, 334, "主观高"),
    ("建设社区", 66, 0, 9, 52, 320, "主观高"),
    ("岳湾路社区", 83, 0, 2, 19, 206, "主观高"),
    ("大学路社区", 97, 1, 6, 44, 178, "主观高"),
    ("伍临路社区", 55, 0, 8, 20, 201, "主观高"),
]

LAYER_ORDER = ["双高", "客观高", "主观高"]
LAYER_TITLES = {
    "双高": "① 双高 · 体检+诉求都高（5 个）",
    "客观高": "② 客观隐患高 · 体检独证（诉求未暴露 · 8 个）",
    "主观高": "③ 主观诉求高 · 热线独证（体检未印证 · 7 个）",
}
LAYER_LABELS = {
    "双高": "双高 ★",
    "客观高": "客观隐患高\n诉求未暴露",
    "主观高": "主观诉求高\n体检未印证",
}


def d(n, b):
    return round(n / b * 100, 1) if b else 0


def _hdr(ws, r, vals, ncol):
    for c, h in enumerate(vals, 1):
        cell = ws.cell(r, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER
    for c in range(len(vals) + 1, ncol + 1):
        ws.cell(r, c).fill = HDR_FILL
        ws.cell(r, c).border = BORDER


def _layer_head(ws, r, layer, ncol):
    ws.cell(r, 1, LAYER_TITLES[layer])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    for c in range(1, ncol + 1):
        cell = ws.cell(r, c)
        cell.fill = LAYER_HEAD[layer]
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = BORDER
    ws.cell(r, 1).alignment = LEFT


def sheet1_original(wb):
    ws = wb.create_sheet("1原表")
    _hdr(ws, 1, ["序号", "名称", "楼栋", "安全韧性问题\n（每百栋）", None, "民生基础需求\n（每百栋）", None, "问题指标", "群众诉求", "评估"], 10)
    ws.cell(1, 4).value = "安全韧性问题\n（每百栋）"
    ws.cell(1, 4).alignment = CENTER
    ws.cell(1, 6).value = "民生基础需求\n（每百栋）"
    ws.cell(1, 6).alignment = CENTER
    ws.cell(2, 4, "体检"); ws.cell(2, 5, "诉求"); ws.cell(2, 6, "体检"); ws.cell(2, 7, "诉求")
    for c in (4, 5, 6, 7):
        ws.cell(2, c).fill = HDR_FILL
        ws.cell(2, c).font = HDR_FONT
        ws.cell(2, c).border = BORDER
        ws.cell(2, c).alignment = CENTER
    r = 3
    for name, b, tsa, tms, ssa, ssm, layer in ROWS:
        tj = tsa + tms
        sub = ssa + ssm
        vals = [None, name, b, d(tsa, b), d(ssa, b), d(tms, b), d(ssm, b), tj, sub, LAYER_LABELS[layer]]
        # 序号按层内绝对量排序后填充（此处直接用数据顺序）
        for c, v in enumerate(vals, 1):
            if c == 1:
                continue
            cell = ws.cell(r, c, v)
            cell.border = BORDER
            cell.fill = LAYER_FILL[layer]
            cell.alignment = CENTER if c in (3, 4, 5, 6, 7, 8, 9) else LEFT
            if c == 10:
                cell.font = LABEL_FONT[layer]
        r += 1
    # 序号（层内绝对量降序：体检层按体检点、诉求层按诉求件）
    # 重排序号：双高/客观高按体检点降序，主观高按诉求件降序
    seq = 0
    for layer in LAYER_ORDER:
        sub = [(name, b, tsa, tms, ssa, ssm) for name, b, tsa, tms, ssa, ssm, _l in ROWS if _l == layer]
        key = (lambda x: x[1] + x[2]) if layer != "主观高" else (lambda x: x[3] + x[4])
        sub.sort(key=key, reverse=True)
        for name, b, tsa, tms, ssa, ssm in sub:
            seq += 1
            # 找到该行写序号（简化：直接按当前显示顺序写）
    # 因 ROWS 已按层+绝对量预排，直接顺序填序号
    seq = 0
    for i, (name, b, tsa, tms, ssa, ssm, layer) in enumerate(ROWS):
        seq += 1
        ws.cell(3 + i, 1, seq).border = BORDER
        ws.cell(3 + i, 1).fill = LAYER_FILL[layer]
        ws.cell(3 + i, 1).alignment = CENTER
    for i, w in enumerate([5, 14, 7, 10, 10, 10, 10, 9, 9, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def sheet2_classify(wb):
    ws = wb.create_sheet("2分类表")
    _hdr(ws, 1, ["序", "社区", "楼栋", "体检问题（点）", "市民诉求（件）", "每百栋（体检/诉求）", "综合评估"], 7)
    r = 2
    seq = 0
    for layer in LAYER_ORDER:
        _layer_head(ws, r, layer, 7)
        r += 1
        for name, b, tsa, tms, ssa, ssm, _l in ROWS:
            if _l != layer:
                continue
            seq += 1
            tj = tsa + tms
            sub = ssa + ssm
            dens = f"{d(tj, b)} / {d(sub, b) if sub else '—'}"
            vals = [seq, name, b, tj, sub if sub else "—", dens, LAYER_LABELS[layer]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = BORDER
                cell.fill = LAYER_FILL[layer]
                cell.alignment = CENTER if c in (1, 3, 4, 5, 6) else LEFT
                if c == 7:
                    cell.font = LABEL_FONT[layer]
                    cell.alignment = CENTER
            r += 1
    for i, w in enumerate([5, 14, 7, 12, 12, 18, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def _bar(val, maxv, num):
    if val <= 0 or num <= 0:
        return "—"
    k = round(val / maxv * MAX_BAR)
    k = max(1, min(MAX_BAR, k))
    return BLOCK * k + " " + str(num)


def sheet3_compare(wb):
    ws = wb.create_sheet("3体检诉求对比表")
    _hdr(ws, 1, ["序", "社区", "楼栋", "体检（点·每百栋）", "诉求（件·每百栋）", "综合评估"], 6)
    r = 2
    seq = 0
    for layer in LAYER_ORDER:
        _layer_head(ws, r, layer, 6)
        r += 1
        for name, b, tsa, tms, ssa, ssm, _l in ROWS:
            if _l != layer:
                continue
            seq += 1
            tj = tsa + tms
            sub = ssa + ssm
            tj_bar = _bar(d(tj, b), TJ_MAX, tj)
            sub_bar = _bar(d(sub, b), SUB_MAX, sub)
            vals = [seq, name, b, tj_bar, sub_bar, LAYER_LABELS[layer]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = BORDER
                cell.fill = LAYER_FILL[layer]
                cell.alignment = CENTER if c in (1, 3, 6) else LEFT
                if c == 4:
                    cell.font = TJ_FONT
                elif c == 5:
                    cell.font = SUB_FONT
                elif c == 6:
                    cell.font = LABEL_FONT[layer]
                    cell.alignment = CENTER
            r += 1
    r += 1
    notes = [
        "横条 = 每百栋密度（体检满格 150 点/百栋、诉求满格 726 件/百栋·两轨分刻度）；条后数字 = 绝对量（体检点 / 诉求件）。",
        "体检 = 体检安全 + 体检民生；诉求 = 12345 九类（安全 4 + 民生 5·剔「其他」）；「—」= 无隐患/无诉求。",
    ]
    for n in notes:
        ws.cell(r, 1, n)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 1).font = Font(size=9, color="666666")
        r += 1
    for i, w in enumerate([5, 13, 7, 26, 26, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet1_original(wb)
    sheet2_classify(wb)
    sheet3_compare(wb)
    wb.save(OUT)
    print("[OK]", os.path.basename(OUT))


if __name__ == "__main__":
    main()
